import streamlit as st
import pandas as pd
import os
import sqlite3
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER

# --- CONFIG (Notion Peek Look) ---
st.set_page_config(page_title="Shift Log Input", layout="centered", page_icon="📝")

# --- UI STYLING ---
st.markdown("""
<style>
    .stApp { background-color: #F7F7F5; }
    .block-container {
        background-color: #FFFFFF;
        padding: 3rem;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        max-width: 800px;
        border: 1px solid #E0E0E0;
    }
    .main-header {
        background-color: #FFD100; 
        padding: 1.5rem; 
        border-bottom: 3px solid black; 
        margin: -3rem -3rem 2rem -3rem;
        border-radius: 12px 12px 0 0;
    }
    .stButton button {
        border-radius: 8px;
    }
</style>
""", unsafe_allow_html=True)

# --- FILE PATHS ---
DATA_DIR = "data"
DB_FILE = os.path.join(DATA_DIR, "shift_logs.db")
ASSETS_FILE = os.path.join(DATA_DIR, "assets.csv")
SPARES_FILE = os.path.join(DATA_DIR, "spares.csv")

# --- DATABASE INIT ---
def init_db():
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    # Tables
    c.execute('''CREATE TABLE IF NOT EXISTS reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, shift TEXT, engineer TEXT, 
        second_engineer TEXT, urgent_notes TEXT, 
        radios_charged BOOLEAN, phones_working BOOLEAN, 
        keys_handed BOOLEAN, safety_check BOOLEAN,
        submitted_at TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS reactives (
        id INTEGER PRIMARY KEY AUTOINCREMENT, report_id INTEGER, asset TEXT, 
        time_called TEXT, time_back TEXT, fault TEXT, engineers INTEGER, 
        description TEXT, downtime REAL, 
        FOREIGN KEY(report_id) REFERENCES reports(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS ppms (
        id INTEGER PRIMARY KEY AUTOINCREMENT, report_id INTEGER, asset TEXT, 
        ppm_id TEXT, status TEXT, comments TEXT, 
        FOREIGN KEY(report_id) REFERENCES reports(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS other_tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT, report_id INTEGER, category TEXT, 
        status TEXT, comments TEXT, 
        FOREIGN KEY(report_id) REFERENCES reports(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS spares_used (
        id INTEGER PRIMARY KEY AUTOINCREMENT, report_id INTEGER, art_number TEXT, 
        description TEXT, location TEXT, quantity INTEGER, decision TEXT, category_code INTEGER,
        FOREIGN KEY(report_id) REFERENCES reports(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS spares_inventory (
        art_number TEXT PRIMARY KEY, description TEXT, location TEXT, 
        category TEXT, stock_level INTEGER, min_stock_level INTEGER
    )''')

    # Smart Migration
    try: c.execute("ALTER TABLE reports ADD COLUMN second_engineer TEXT")
    except: pass
    try: c.execute("ALTER TABLE reports ADD COLUMN keys_handed BOOLEAN")
    except: pass
    try: c.execute("ALTER TABLE reports ADD COLUMN safety_check BOOLEAN")
    except: pass
    try: c.execute("ALTER TABLE spares_used ADD COLUMN category_code INTEGER")
    except: pass
    try: c.execute("ALTER TABLE spares_used ADD COLUMN location TEXT") 
    except: pass

    conn.commit()
    conn.close()

# --- FUNCTION: FORCE RELOAD SPARES ---
def reload_spares_from_csv():
    if not os.path.exists(SPARES_FILE):
        st.sidebar.error(f"❌ File not found: {SPARES_FILE}")
        return
    
    try:
        df = pd.read_csv(SPARES_FILE)
        df.columns = df.columns.str.strip()
        
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("DELETE FROM spares_inventory")
        
        count = 0
        duplicates = 0
        
        for _, row in df.iterrows():
            p_num = str(row.get('art_number', row.get('Part Number', 'UNKNOWN')))
            desc = str(row.get('Description', row.get('description', '')))
            loc = str(row.get('Location', row.get('location', '')))
            
            if p_num == 'UNKNOWN' or p_num == 'nan': continue

            try:
                c.execute("""
                    INSERT OR REPLACE INTO spares_inventory 
                    (art_number, description, location, category, stock_level, min_stock_level) 
                    VALUES (?,?,?,?,?,?)
                """, (p_num, desc, loc, 'General', 10, 2))
                count += 1
            except sqlite3.IntegrityError: duplicates += 1
            
        conn.commit()
        conn.close()
        st.sidebar.success(f"✅ Loaded {count} spares! (Handled {duplicates} duplicates)")
        
    except Exception as e:
        st.sidebar.error(f"Error reading CSV: {e}")

# --- EXCEL ENGINE ---
def generate_excel_report(shift_data, reactives, ppms, other_tasks, spares):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift Report"
    
    # Styles
    ssi_yellow = PatternFill(start_color="FFD100", end_color="FFD100", fill_type="solid")
    blue_header = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    white_font = Font(color="FFFFFF", bold=True)
    
    # Header
    ws['A1'] = "Engineer:"; ws['B1'] = shift_data['engineer']
    ws['C1'] = shift_data.get('second_engineer', '')
    ws.merge_cells('E1:L6')
    ws['E1'] = f"HANDOVER NOTES:\n{shift_data['urgent_notes']}"
    ws['E1'].alignment = Alignment(wrap_text=True, vertical='top')
    
    checks = [
        ("RADIO CHECK", shift_data['radios_charged']), 
        ("PHONES HANDED", shift_data['phones_working']),
        ("KEYS HANDED", shift_data['keys_handed']),
        ("SAFETY OK", shift_data['safety_check'])
    ]
    for i, (lbl, val) in enumerate(checks):
        ws[f'A{3+i}'] = lbl
        ws[f'B{3+i}'] = "☑" if val else "☐"
        ws[f'A{3+i}'].border = thin_border
        ws[f'B{3+i}'].border = thin_border

    # Sections
    row = 8
    # Reactives
    ws[f'A{row}'] = "Reactive Tasks"; ws[f'A{row}'].fill = ssi_yellow; ws[f'A{row}'].font = bold_font
    ws.merge_cells(f'A{row}:L{row}'); row += 1
    headers = ['Time Called', 'Time Back', 'Asset', 'Fault', 'Comment', 'Engs', 'Man Hrs', 'DT Hrs']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = PatternFill("E0E0E0", fill_type="solid"); cell.border = thin_border
    row += 1
    for r in reactives:
        dt = r['Downtime (min)']/60
        vals = [r['Time Called'], r['Time Back'], r['Asset'], r['Fault'], r['Description'], r['Engineers'], round(dt*r['Engineers'],2), round(dt,2)]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v).border = thin_border
        row += 1
    
    # PPMs
    row += 2
    ws[f'A{row}'] = "PPM Tasks"; ws[f'A{row}'].fill = ssi_yellow; ws[f'A{row}'].font = bold_font
    ws.merge_cells(f'A{row}:D{row}'); row += 1
    for p in ppms:
        ws.cell(row=row, column=1, value=p['Asset']).border = thin_border
        ws.cell(row=row, column=2, value=p['Status']).border = thin_border
        ws.cell(row=row, column=3, value=p['Comments']).border = thin_border
        row += 1

    # Spares
    row += 2
    ws[f'A{row}'] = "Spares Used"; ws[f'A{row}'].fill = ssi_yellow; ws[f'A{row}'].font = bold_font
    ws.merge_cells(f'A{row}:L{row}'); row += 1
    s_heads = ['ART #', 'LOCATION', 'DESC', 'CATEGORY', 'REASON']
    for c, h in enumerate(s_heads, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = blue_header; cell.font = white_font
    
    row += 1
    s_heads2 = ['QTY', 'NAME', 'DATE', 'DECISION']
    for c, h in enumerate(s_heads2, 6):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = blue_header; cell.font = white_font

    row += 1
    for s in spares:
        vals = [s['ART #'], s['Location'], s['Description'], s.get('Category Code', ''), "Wear Part"]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v).border = thin_border
        vals2 = [s['Quantity'], shift_data['engineer'], shift_data['date'], s['Decision']]
        for c, v in enumerate(vals2, 6):
            ws.cell(row=row+1, column=c, value=v).border = thin_border
        row += 2

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- PDF ENGINE ---
def generate_pdf_report(shift_data, reactives, ppms, other_tasks, spares, rid):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('T', parent=styles['Heading1'], backColor=colors.HexColor('#FFD100'), alignment=TA_CENTER, borderPadding=10)
    
    # Page 1
    elements.append(Paragraph(f"SHIFT REPORT #{rid}", title_style))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Engineer:</b> {shift_data['engineer']} | <b>Date:</b> {shift_data['date']}", styles['Normal']))
    elements.append(Paragraph(f"<b>Handover Notes:</b> {shift_data['urgent_notes']}", styles['Normal']))
    elements.append(PageBreak())
    
    # Page 2
    elements.append(Paragraph("Reactive Tasks", title_style))
    if reactives:
        data = [['Asset', 'Fault', 'Time', 'DT(m)']] + [[r['Asset'], r['Fault'], r['Time Called'], r['Downtime (min)']] for r in reactives]
        t = Table(data, colWidths=[2*inch, 2*inch, 1*inch, 1*inch])
        t.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black), ('BACKGROUND', (0,0), (-1,0), colors.lightgrey)]))
        elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

# --- MASTER DATA ---
def get_master_data():
    if os.path.exists(ASSETS_FILE):
        try:
            df_assets = pd.read_csv(ASSETS_FILE)
            assets = df_assets['Asset Name'].dropna().unique().tolist()
        except: assets = ["Conveyor 1", "Wrapper", "Palletiser"]
    else: assets = ["Conveyor 1", "Wrapper", "Palletiser"]

    conn = sqlite3.connect(DB_FILE)
    try:
        df_spares = pd.read_sql("SELECT * FROM spares_inventory", conn)
        if not df_spares.empty:
            spares_list = (df_spares['art_number'] + " - " + df_spares['description']).tolist()
        else:
            spares_list = []
            df_spares = pd.DataFrame()
    except:
        spares_list = []
        df_spares = pd.DataFrame()
    conn.close()
    return sorted(assets), sorted(spares_list), df_spares

# --- SESSION STATE ---
def init_session_state():
    if 'reactives' not in st.session_state: st.session_state.reactives = []
    if 'ppms' not in st.session_state: st.session_state.ppms = []
    if 'other_tasks' not in st.session_state: st.session_state.other_tasks = []
    if 'spares' not in st.session_state: st.session_state.spares = []

# --- MAIN APP ---
init_db()
init_session_state()

# SIDEBAR RELOAD
with st.sidebar:
    st.header("⚙️ Settings")
    if st.button("🔄 Force Reload Spares from CSV"):
        reload_spares_from_csv()

ASSETS, SPARES_OPT, DF_SPARES = get_master_data()

st.markdown("""<div class="main-header"><h1>📝 ENTER SHIFT LOG</h1></div>""", unsafe_allow_html=True)

# 1. INFO
c1, c2, c3 = st.columns(3)
s_date = c1.date_input("Date", datetime.now())
s_shift = c2.selectbox("Shift", ["Day", "Night"])
s_eng = c3.selectbox("Lead Engineer", ["Chris McGhee", "Alf Fidoe", "Mike Wilson", "Sarah Jones"])
s_sec_eng = st.text_input("Second Engineer (Optional)")

# 2. REACTIVES
with st.expander("⚠️ Reactive Tasks", expanded=True):
    c1, c2, c3, c4 = st.columns(4)
    r_asset = c1.selectbox("Asset", ASSETS)
    r_fault = c2.text_input("Fault Type")
    r_called = c3.time_input("Call Time")
    r_back = c4.time_input("Back Time")
    r_desc = st.text_area("Description / Action")
    
    if st.button("➕ Add Reactive"):
        dt = (datetime.combine(datetime.today(), r_back) - datetime.combine(datetime.today(), r_called)).total_seconds() / 60
        if dt < 0: dt += 1440
        st.session_state.reactives.append({
            "Asset": r_asset, "Fault": r_fault, "Time Called": str(r_called), 
            "Time Back": str(r_back), "Downtime (min)": round(dt, 1), 
            "Engineers": 2 if s_sec_eng else 1, "Description": r_desc
        })
    if st.session_state.reactives: st.dataframe(pd.DataFrame(st.session_state.reactives), use_container_width=True)

# 3. SPARES
with st.expander("📦 Spares Used"):
    c1, c2 = st.columns([3, 1])
    sp_item = c1.selectbox("Search Part", SPARES_OPT)
    
    c3, c4, c5 = st.columns(3)
    sp_qty = c3.number_input("Qty", 1, 100, 1)
    sp_cat = c4.selectbox("Category", [1, 2, 3, 4, 5, 6, 7, 8, 9])
    sp_dec = c5.selectbox("Decision", ["Disposed", "Quarantined"])
    
    if st.button("➕ Add Spare"):
        if " - " in sp_item and not DF_SPARES.empty:
            art = sp_item.split(' - ')[0]
            row = DF_SPARES[DF_SPARES['art_number'] == art]
            if not row.empty:
                st.session_state.spares.append({
                    "ART #": art, 
                    "Description": row.iloc[0]['description'], 
                    "Location": row.iloc[0]['location'],
                    "Quantity": sp_qty, 
                    "Category Code": sp_cat, 
                    "Decision": sp_dec
                })
    if st.session_state.spares: 
        st.dataframe(pd.DataFrame(st.session_state.spares), use_container_width=True)

# 4. PPM / OTHER (FIXED: Now displays the table!)
with st.expander("🛠️ PPM & Other Tasks"):
    c1, c2 = st.columns(2)
    p_asset = c1.selectbox("PPM Asset", ASSETS, key='ppm')
    p_comm = c2.text_input("PPM Comment", key='ppm_comm')
    
    if st.button("➕ Add PPM"):
        st.session_state.ppms.append({"Asset": p_asset, "Status": "Complete", "Comments": p_comm})
        st.rerun() # Forces the app to refresh and show the new entry immediately

    # THIS WAS MISSING:
    if st.session_state.ppms:
        st.markdown("**📋 PPM List**")
        st.table(pd.DataFrame(st.session_state.ppms))

# 5. SUBMIT
st.divider()
urgent = st.text_area("📝 Handover Notes")
c1, c2, c3, c4 = st.columns(4)
chk_rad = c1.checkbox("Radios Charged")
chk_phn = c2.checkbox("Phones Handed")
chk_key = c3.checkbox("Keys Handed")
chk_safe = c4.checkbox("Safety Check")

if st.button("💾 SAVE REPORT", type="primary"):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT INTO reports (date, shift, engineer, second_engineer, urgent_notes, radios_charged, phones_working, keys_handed, safety_check, submitted_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
              (str(s_date), s_shift, s_eng, s_sec_eng, urgent, chk_rad, chk_phn, chk_key, chk_safe, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    rid = c.lastrowid
    
    for r in st.session_state.reactives:
        c.execute("INSERT INTO reactives (report_id, asset, time_called, time_back, fault, engineers, description, downtime) VALUES (?,?,?,?,?,?,?,?)",
                 (rid, r['Asset'], r['Time Called'], r['Time Back'], r['Fault'], r['Engineers'], r['Description'], r['Downtime (min)']))
    
    for s in st.session_state.spares:
        c.execute("INSERT INTO spares_used (report_id, art_number, description, location, quantity, decision, category_code) VALUES (?,?,?,?,?,?,?)",
                 (rid, s['ART #'], s['Description'], s['Location'], s['Quantity'], s['Decision'], s['Category Code']))
        c.execute("UPDATE spares_inventory SET stock_level = stock_level - ? WHERE art_number = ?", (s['Quantity'], s['ART #']))
        
    conn.commit()
    conn.close()
    
    r_data = {"date": str(s_date), "shift": s_shift, "engineer": s_eng, "second_engineer": s_sec_eng, "urgent_notes": urgent, 
              "radios_charged": chk_rad, "phones_working": chk_phn, "keys_handed": chk_key, "safety_check": chk_safe}
    
    xls = generate_excel_report(r_data, st.session_state.reactives, st.session_state.ppms, st.session_state.other_tasks, st.session_state.spares)
    pdf = generate_pdf_report(r_data, st.session_state.reactives, st.session_state.ppms, st.session_state.other_tasks, st.session_state.spares, rid)
    
    folder = os.path.join("data", "reports", str(s_date.year), f"{s_date.month:02d}")
    os.makedirs(folder, exist_ok=True)
    bname = f"ShiftReport_{s_date.strftime('%Y%m%d')}_{s_shift}_{s_eng.replace(' ','')}"
    
    with open(os.path.join(folder, f"{bname}.xlsx"), "wb") as f: f.write(xls.getbuffer())
    with open(os.path.join(folder, f"{bname}.pdf"), "wb") as f: f.write(pdf.getbuffer())
    
    st.success("Report Saved!")
    st.download_button("Download Excel", xls, f"{bname}.xlsx")
    st.download_button("Download PDF", pdf, f"{bname}.pdf")